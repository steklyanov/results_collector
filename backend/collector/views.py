from .models import Catch, Person
from .serializers import NestedCatchesSerializer, CatchSerializer, PersonSerializer

from django.core.files.base import ContentFile
from django.db import connection
from django.http import HttpResponse

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework import generics

import os
import names
import base64
import time
import datetime
from openpyxl import Workbook
from shutil import copy, make_archive, rmtree


class CatchHandler(APIView):
    def post(self, request):
        if "ex_id" and "image" in request.data:
            person, _ = Person.objects.get_or_create(ex_id=request.data["ex_id"],
                                                     defaults={'name': names.get_full_name()})
            image_string = request.data["image"]
            Catch.objects.create(person=person,
                                 image=ContentFile(base64.b64decode(image_string),
                                                   name=str(request.data["ex_id"]) +
                                                        time.strftime("%Y%m%d-%H%M%S") + '.jpg'))

            return Response(status=status.HTTP_201_CREATED)
        else:
            return Response({"error": "no external id or image provided"}, status=status.HTTP_204_NO_CONTENT)


class ListCatchesView(generics.ListAPIView):
    serializer_class = NestedCatchesSerializer

    def get_queryset(self, *args, **kwargs):
        from_time = datetime.datetime.fromtimestamp(int(self.request.query_params.get("from",
                                                                             int(datetime.datetime.now().timestamp()))))
        till_time = datetime.datetime.fromtimestamp(int(self.request.query_params.get("till",
                                                                             int(datetime.datetime.now().timestamp()))))
        cursor = connection.cursor()
        cursor.execute('''select person_id, array_agg(t2.*) as catch_groups from (select * from collector_catch
                        left join "collector_person"
                        on ("collector_catch"."person_id" = "collector_person"."id")
                        where datetime between %s and %s
                        order by collector_catch.datetime, collector_catch.person_id asc ) t2 group by t2.person_id''',
                       (from_time, till_time))
        rows = cursor.fetchall()
        print(rows)
        return rows

    def list(self, request, *args, **kwargs):
        is_report = self.request.query_params.get("report", False)
        if is_report:
            from_time = datetime.datetime.fromtimestamp(int(self.request.query_params.get("from",
                                                                                          int(datetime.datetime.now().timestamp()))))
            till_time = datetime.datetime.fromtimestamp(int(self.request.query_params.get("till",
                                                                                      int(datetime.datetime.now().timestamp()))))
            catches = Catch.objects.filter(datetime__date__gte=from_time, datetime__date__lte=till_time).order_by('datetime')
            print(catches)
            wb = Workbook()
            ws = wb.active
            ws.title = 'Пассажиропоток'
            idx = 1
            for catch in catches:
                ws['A' + str(idx + 1)] = str(catch.person.name)
                ws['B' + str(idx + 1)] = str(catch.person.ex_id)
                ws['C' + str(idx + 1)] = str(catch.datetime)
                idx += 1
            wb.save('documents/sheet.xlsx')
            make_archive("document", 'zip', "documents")
            response = HttpResponse(open("document.zip", 'rb'), content_type='application/zip')
            return response
        return super().list(request, args, kwargs)


class PersonListView(generics.ListAPIView):
    serializer_class = PersonSerializer
    queryset = Person.objects.all()


class SinglePersonView(generics.ListAPIView):
    lookup_url_kwarg = "id"
    serializer_class = CatchSerializer

    def get_queryset(self, *args, **kwargs):
        print(self.request)
        print(self.request.query_params)
        print(args)
        print(kwargs)
        uid = self.kwargs.get(self.lookup_url_kwarg)
        print("UUID", uid)
        rows = Catch.objects.filter(person_id=uid)
        return rows


class GetLastImage(APIView):
    def get(self, request):
        # I DONT NEED CAM_ID SO DEFAULT VALUE TRUE AND NOT FILTERED
        cam_id = self.request.query_params.get("from", True)
        if cam_id:
            prev_issue = Catch.objects.last()
            serializer = CatchSerializer(prev_issue)
            return Response(serializer.data)
        return Response({"error": "Image not found"}, status=400)
